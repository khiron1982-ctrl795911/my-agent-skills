#!/usr/bin/env python3
"""勤怠自動チェック_YYYYMM.xlsx（kintai_agent.py の出力）を読み、
「誰に・何を確認すべきか」の要確認サマリーを Markdown で出力する。

使い方:
  py -X utf8 summarize_check.py <勤怠自動チェック_YYYYMM.xlsx> [--md 出力先.md]

- 各チェックシートの判定列（判定/残業判定/問題）を見て、✅/➖ 以外の行を「要確認」として拾う
- 判定列がないシート（遅刻など、載っている行自体が例外リスト）は名前入り全行を要確認とする
- 出力は (1) シート別サマリー (2) 人別の確認事項リスト（メール下書きの入力に使える形式）
- このスクリプトは読み取り専用。入力ファイルを一切変更しない。
"""
import sys
import io
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

JUDGE_HEADERS = ("判定", "残業判定", "問題")
OK_PREFIXES = ("✅", "➖", "OK", "○")
# 補足情報として拾う列（存在すれば）
DETAIL_HEADERS = ("日付", "区分", "備考", "コメント", "差異", "問題", "超過(分)")
SUMMARY_SHEET_KEYWORDS = ("サマリー",)


def clean(v):
    return str(v).strip() if v is not None else ""


def find_header_row(ws):
    """先頭8行から「名前」を含む行をヘッダーとして探す"""
    for r in range(1, min(9, ws.max_row + 1)):
        vals = [clean(c.value) for c in ws[r]]
        if "名前" in vals:
            return r, vals
    return None, None


def summarize(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_results = []          # (sheet, total_rows, flagged_rows[list of dict])
    person_issues = defaultdict(list)  # 名前 -> ["シート名: 詳細", ...]

    for ws in wb.worksheets:
        if any(k in ws.title for k in SUMMARY_SHEET_KEYWORDS):
            continue
        hdr_row, headers = find_header_row(ws)
        if hdr_row is None:
            continue
        idx = {h: i for i, h in enumerate(headers) if h}
        name_i = idx["名前"]
        judge_i = next((idx[h] for h in JUDGE_HEADERS if h in idx), None)
        detail_is = [(h, idx[h]) for h in DETAIL_HEADERS if h in idx]

        total = 0
        flagged = []
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            name = clean(row[name_i]) if name_i < len(row) else ""
            if not name or name.startswith(("※", "✅")):
                continue
            total += 1
            if judge_i is not None:
                judge = clean(row[judge_i]) if judge_i < len(row) else ""
                is_flag = bool(judge) and not judge.startswith(OK_PREFIXES)
            else:
                # 判定列がないシートは掲載行そのものが例外リスト
                judge = ""
                is_flag = True
            if is_flag:
                details = []
                for h, i in detail_is:
                    v = clean(row[i]) if i < len(row) else ""
                    if v:
                        details.append(f"{h}={v}")
                info = {"名前": name, "判定": judge, "詳細": " / ".join(details)}
                flagged.append(info)
                desc = judge or "要確認"
                if info["詳細"]:
                    desc += f"（{info['詳細']}）"
                person_issues[name].append(f"{ws.title}: {desc}")
        sheet_results.append((ws.title, total, flagged))

    return sheet_results, person_issues


def compact_issues(issues):
    """同一シート・同一判定の指摘が4件以上あれば、日付をまとめて1行に圧縮する"""
    import re as _re
    grouped = defaultdict(list)
    order = []
    for issue in issues:
        m = _re.match(r"^([^:]+): ([^（(]+)", issue)
        key = (m.group(1), m.group(2).strip()) if m else (issue, "")
        if key not in grouped:
            order.append(key)
        grouped[key].append(issue)
    out = []
    for key in order:
        items = grouped[key]
        if len(items) < 4:
            out.extend(items)
            continue
        dates = []
        for it in items:
            dm = _re.search(r"日付=(\d{4}/\d{1,2}/\d{1,2}(?:\([^)]*\))?)", it)
            if dm:
                dates.append(_re.sub(r"^\d{4}/", "", dm.group(1)))
        label = f"{key[0]}: {key[1]} × {len(items)}件"
        if dates:
            label += f"（{', '.join(dates)}）"
        out.append(label)
    return out


def to_markdown(path: Path, sheet_results, person_issues):
    lines = [f"# 勤怠チェック 要確認サマリー", f"", f"入力: `{path.name}`", ""]
    lines.append("## シート別サマリー")
    lines.append("")
    lines.append("| チェック項目 | 対象行数 | 要確認 |")
    lines.append("|---|---:|---:|")
    total_flags = 0
    for title, total, flagged in sheet_results:
        mark = f"**{len(flagged)}**" if flagged else "0"
        lines.append(f"| {title} | {total} | {mark} |")
        total_flags += len(flagged)
    lines.append("")
    lines.append(f"**要確認 合計: {total_flags} 件 / 対象者 {len(person_issues)} 名**")
    lines.append("")

    if person_issues:
        lines.append("## 人別の確認事項（本人・上長への確認用）")
        lines.append("")
        for name in sorted(person_issues, key=lambda n: -len(person_issues[n])):
            issues = compact_issues(person_issues[name])
            lines.append(f"### {name}（{len(person_issues[name])}件）")
            for issue in issues:
                lines.append(f"- {issue}")
            lines.append("")
    else:
        lines.append("✅ 要確認事項はありません。")
        lines.append("")

    lines.append("---")
    lines.append("※ このサマリーは機械判定です。給与に影響する項目（代休付与差異・通勤費差異・残業差異）は")
    lines.append("　 必ず原本Excelの該当シートを目視確認してから連絡・修正してください。")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="勤怠自動チェック_YYYYMM.xlsx のパス")
    ap.add_argument("--md", help="Markdownの出力先パス（省略時は入力と同じフォルダに 確認サマリー_YYYYMM.md）")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"❌ ファイルが見つかりません: {path}")
        sys.exit(1)

    sheet_results, person_issues = summarize(path)
    md = to_markdown(path, sheet_results, person_issues)

    out = Path(args.md) if args.md else path.with_name(
        path.stem.replace("勤怠自動チェック", "確認サマリー") + ".md")
    out.write_text(md, encoding="utf-8")
    print(md)
    print(f"\n📄 保存先: {out}")


if __name__ == "__main__":
    main()
