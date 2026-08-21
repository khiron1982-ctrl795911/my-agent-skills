from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

BASE26 = [
    "雇用区分コード", "従業員コード", "名前", "休憩時間", "総労働時間(独自)", "所定時間(独自)",
    "総残業時間(SI)", "所定外労働時間", "法定外労働時間", "所定休日勤務時間", "法定休日勤務時間", "深夜時間",
    "総出勤日数", "所定休出日数", "法定休日出勤日数", "有休取得日数", "有休時間休取得時間",
    "特別休暇(有給)", "特別休暇(無給)", "欠勤取得日数", "早退回数", "早退時間", "遅刻回数", "遅刻時間",
    "代休取得日数", "代休付与日数",
]

FLEX35 = [
    "雇用区分コード", "従業員コード", "名前", "休憩時間", "総労働時間(独自)", "所定時間(独自)",
    "休暇取得日数", "休暇取得時間", "所定時間\n(休暇取得時間合算)", "所定内不足時間",
    "所定外労働時間\n(不足分調整)", "休業日数", "総残業時間(確定)", "所定外労働時間\n(調整)",
    "法定外労働時間\n(調整)", "総残業時間(SI)", "所定外労働時間", "法定外労働時間", "所定休日勤務時間",
    "法定休日勤務時間", "深夜時間", "総出勤日数", "所定休出日数", "法定休日出勤日数", "有休取得日数",
    "有休取得時間", "特別休暇(有給)", "特別休暇(無給)", "欠勤取得日数", "早退回数", "早退時間", "遅刻回数",
    "遅刻時間", "代休取得日数", "代休付与日数",
]

SIXTY6 = ["雇用区分コード", "従業員コード", "名前", "法定外労働時間", "所定休日勤務時間", "60時間超"]
EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
OUTPUT_NAME_TOKENS = ("5区分", "フレックス", "60時間", "一般アルバイト", "振替休日", "振替出勤", "振休", "振出")

@dataclass
class SheetData:
    path: Path
    sheet: str
    headers: list[str]
    rows: list[list[Any]]


def norm_header(value: Any) -> str:
    return str(value or "").replace("\r\n", "\n").strip()


def header_key(value: str) -> str:
    return re.sub(r"\s+", "", value.replace("\n", ""))


def clean_path(text: str) -> Path:
    return Path(text.strip().strip('"').strip("'"))


def discover(paths: list[Path]) -> list[Path]:
    files: list[Path] = []
    for p in paths:
        if p.is_dir():
            files.extend(
                x for x in p.rglob("*")
                if x.suffix.lower() in EXCEL_EXTS
                and not x.name.startswith("~$")
                and any(t in x.name for t in OUTPUT_NAME_TOKENS)
            )
        elif p.is_file() and p.suffix.lower() in EXCEL_EXTS and not p.name.startswith("~$"):
            files.append(p)
            files.extend(
                x for x in p.parent.glob("*")
                if x.suffix.lower() in EXCEL_EXTS
                and not x.name.startswith("~$")
                and any(t in x.name for t in OUTPUT_NAME_TOKENS)
            )
    return sorted(set(files), key=lambda x: str(x))


def find_header(values: list[tuple[Any, ...]]) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(values, start=1):
        headers = [norm_header(v) for v in row]
        if "従業員コード" in headers and "名前" in headers:
            return idx, headers
    return None


def read_matching_sheets(path: Path) -> list[SheetData]:
    out: list[SheetData] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return out
    for ws in wb.worksheets:
        first_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True))
        found = find_header(first_rows)
        if not found:
            continue
        hr, headers = found
        rows: list[list[Any]] = []
        for row in ws.iter_rows(min_row=hr + 1, max_row=min(ws.max_row, hr + 2500), values_only=True):
            vals = list(row)
            if any(v not in (None, "") for v in vals):
                rows.append(vals)
        out.append(SheetData(path, ws.title, headers, rows))
    return out


def header_positions(headers: list[str]) -> dict[str, list[int]]:
    pos: dict[str, list[int]] = {}
    for i, h in enumerate(headers):
        pos.setdefault(header_key(h), []).append(i)
    return pos


def has_all(headers: list[str], required: list[str]) -> bool:
    keys = set(header_key(h) for h in headers)
    return all(header_key(r) in keys for r in required)


def select_columns(
    data: SheetData,
    target_headers: list[str],
    use_last_duplicate: bool = False,
    duplicate_last_headers: set[str] | None = None,
) -> list[list[Any]]:
    positions = header_positions(data.headers)
    selected: list[int] = []
    missing: list[str] = []
    duplicate_last_keys = {header_key(h) for h in duplicate_last_headers or set()}
    for h in target_headers:
        matches = positions.get(header_key(h), [])
        if not matches:
            missing.append(h)
            selected.append(-1)
        else:
            use_last = use_last_duplicate or header_key(h) in duplicate_last_keys
            selected.append(matches[-1] if use_last else matches[0])
    if missing:
        raise ValueError(f"missing columns in {data.path.name}: {missing}")
    return [[row[i] if i < len(row) else None for i in selected] for row in data.rows]


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def classify(data: SheetData) -> str | None:
    name = str(data.path)
    if has_all(data.headers, SIXTY6) or "60時間" in name:
        return "60h"
    if "フレックス" in name and has_all(data.headers, FLEX35):
        return "flex"
    if "一般アルバイト" in name and has_all(data.headers, BASE26):
        return "parttime"
    if "5区分" in name and has_all(data.headers, BASE26):
        return "five"
    return None



def choose_flex_base(candidates: list[SheetData]) -> SheetData | None:
    filtered = [c for c in candidates if "フレックス" in str(c.path) and has_all(c.headers, BASE26)]
    if not filtered:
        return None
    return sorted(filtered, key=lambda d: (2 if "秦作成" in str(d.path) else 1, len(d.rows)), reverse=True)[0]


def build_unadjusted_flex_from_base(data: SheetData) -> list[list[Any]]:
    base_rows = select_columns(data, BASE26, use_last_duplicate=False)
    out_rows: list[list[Any]] = []
    for r in base_rows:
        # This fallback preserves raw attendance values and leaves flex-specific
        # adjustment columns blank because monthly standard-hours logic is not
        # proven from a 26-column source alone.
        out_rows.append([
            r[0], r[1], r[2], r[3], r[4], r[5],
            None, None, None, None, None, None, None, None, None,
            r[6], r[7], r[8], r[9], r[10], r[11], r[12], r[13], r[14], r[15], r[16],
            r[17], r[18], r[19], r[20], r[21], r[22], r[23], r[24], r[25],
        ])
    return out_rows
def choose_best(candidates: list[SheetData], kind: str) -> SheetData | None:
    filtered = [c for c in candidates if classify(c) == kind]
    if not filtered:
        return None
    def score(d: SheetData) -> tuple[int, int, int]:
        p = str(d.path)
        return (
            5 if "アックス送付済" in p else 4 if "秦作成" in p else 3 if "みこしば" in p else 2 if "菊池" in p else 1,
            2 if "_3加算用" in p else 1,
            len(d.rows),
        )
    return sorted(filtered, key=score, reverse=True)[0]


def parse_period_from_sheet(sheet: str) -> tuple[int, int] | None:
    m = re.search(r"(20\d{2})(\d{2})\d{2}[_-](20\d{2})(\d{2})\d{2}", sheet)
    if not m:
        return None
    y, month = int(m.group(1)), int(m.group(2))
    return (y + 1, 1) if month == 12 else (y, month + 1)


def payroll_label(payroll_month: str | None, sheets: list[SheetData]) -> str:
    if payroll_month:
        y, m = payroll_month.split("-")
        return f"{int(y)}年{int(m)}月給与"
    for s in sheets:
        parsed = parse_period_from_sheet(s.sheet)
        if parsed:
            return f"{parsed[0]}年{parsed[1]}月給与"
    return "対象月給与"


def find_transfer_files(files: list[Path]) -> list[Path]:
    return [p for p in files if re.search(r"振替休日|振替出勤|振休者|振出", p.name)]


def main() -> int:
    parser = argparse.ArgumentParser(description="Build payroll attendance processed workbooks from raw attendance paths.")
    parser.add_argument("paths", nargs="+", help="Input Excel files or folders. Quoted Windows paths are accepted.")
    parser.add_argument("--out", default=None, help="Output folder. Default: <first input>/加工済み勤怠データ")
    parser.add_argument("--payroll-month", default=None, help="Payroll month in YYYY-MM, e.g. 2025-10.")
    args = parser.parse_args()

    inputs = [clean_path(x) for x in args.paths]
    files = discover(inputs)
    if not files:
        raise SystemExit("No candidate Excel files found. Expected filenames containing 5区分, フレックス, 60時間, 一般アルバイト, or 振替/振休.")

    sheets: list[SheetData] = []
    for f in files:
        sheets.extend(read_matching_sheets(f))
    if not sheets:
        raise SystemExit("No attendance-like sheets found. Expected headers include 従業員コード and 名前.")

    label = payroll_label(args.payroll_month, sheets)
    first = inputs[0]
    out_dir = Path(args.out) if args.out else (first if first.is_dir() else first.parent) / "加工済み勤怠データ"
    out_dir.mkdir(parents=True, exist_ok=True)

    report: dict[str, Any] = {"input_files": [str(f) for f in files], "outputs": {}, "warnings": []}

    five = choose_best(sheets, "five")
    if five:
        out = out_dir / f"【{label}】5区分勤怠データ.xlsx"
        rows = select_columns(five, BASE26)
        write_workbook(out, five.sheet, BASE26, rows)
        report["outputs"]["5区分勤怠データ"] = {"path": str(out), "source": str(five.path), "sheet": five.sheet}
    else:
        report["warnings"].append("5区分勤怠データの元シートが見つかりません。")

    flex = choose_best(sheets, "flex")
    if flex:
        out = out_dir / f"【{label}】フレックス勤務者勤怠データ.xlsx"
        rows = select_columns(flex, FLEX35)
        write_workbook(out, flex.sheet, FLEX35, rows)
        report["outputs"]["フレックス勤務者勤怠データ"] = {"path": str(out), "source": str(flex.path), "sheet": flex.sheet}
    else:
        flex_base = choose_flex_base(sheets)
        if flex_base:
            out = out_dir / f"【{label}】フレックス勤務者勤怠データ.xlsx"
            rows = build_unadjusted_flex_from_base(flex_base)
            write_workbook(out, flex_base.sheet, FLEX35, rows)
            report["outputs"]["フレックス勤務者勤怠データ"] = {"path": str(out), "source": str(flex_base.path), "sheet": flex_base.sheet, "status": "26列元データから仮整形"}
            report["warnings"].append("35列構成のフレックス勤務者勤怠データが見つからないため、26列フレックス元データから仮整形しました。休暇取得日数、所定内不足時間、調整後残業列は要確認です。")
        else:
            report["warnings"].append("フレックス勤務者勤怠データの元シートが見つかりません。")

    sixty = choose_best(sheets, "60h")
    if sixty:
        out = out_dir / f"【{label}】60時間超過対象者(正社員&フレックス).xlsx"
        rows = select_columns(sixty, SIXTY6)
        write_workbook(out, sixty.sheet, SIXTY6, rows)
        report["outputs"]["60時間超過対象者"] = {"path": str(out), "source": str(sixty.path), "sheet": sixty.sheet}
    else:
        report["warnings"].append("60時間超過対象者の元シートが見つかりません。")

    parttime = choose_best(sheets, "parttime")
    if parttime:
        out = out_dir / f"【{label}】一般アルバイト勤怠データ.xlsx"
        rows = select_columns(parttime, BASE26)
        write_workbook(out, parttime.sheet, BASE26, rows)
        report["outputs"]["一般アルバイト勤怠データ"] = {"path": str(out), "source": str(parttime.path), "sheet": parttime.sheet}
    else:
        report["warnings"].append("一般アルバイト勤怠データの元シートが見つかりません。")

    copied = []
    for p in find_transfer_files(files):
        target = out_dir / p.name
        if p.resolve() != target.resolve():
            shutil.copy2(p, target)
        copied.append(str(target))
    if copied:
        report["outputs"]["振替休日/振替出勤データ一覧"] = copied

    report_path = out_dir / "attendance_processing_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["outputs"] else 2


if __name__ == "__main__":
    raise SystemExit(main())





